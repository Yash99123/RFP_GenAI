"""
EITACIES RFP GenAI — Weekly Excel Report Generator
Pulls data from PostgreSQL and generates a styled 3-sheet Excel file.
"""
import psycopg2
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent))
from dotenv import load_dotenv
load_dotenv(Path(__file__).parent.parent.parent / ".env", override=True)

PROJECT_DIR = Path(__file__).parent.parent.parent
OUTPUT_DIR = PROJECT_DIR / "exports" / "weekly"


def get_db():
    return psycopg2.connect(os.environ["DATABASE_URL"])


def pull_records():
    """Pull this week's tenders from PostgreSQL."""
    conn = get_db()
    import pandas as pd
    df = pd.read_sql("""
        SELECT
            solicitation_id, name, source_portal, end_client, priority, score,
            closing_date, questions_deadline, first_seen, country, office, sector,
            contract_type, project_value_usd, project_value_cad,
            details_url, description, notes, flags, red_flags, status
        FROM tenders
        WHERE first_seen >= NOW() - INTERVAL '7 days'
        ORDER BY score DESC NULLS LAST, closing_date ASC NULLS LAST
    """, conn)
    conn.close()
    return df


def generate_excel(df):
    """Generate styled Excel report from DataFrame."""
    wb = Workbook()

    # Styling
    header_fill = PatternFill("solid", fgColor="1A1A18")
    header_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    cell_font = Font(size=10, name="Calibri")
    border = Border(
        bottom=Side(style="thin", color="E5E3DC"),
    )

    tier_colors = {
        "Strategic":          PatternFill("solid", fgColor="FEE2E2"),
        "Tier 1 - Pursue":    PatternFill("solid", fgColor="FFF7ED"),
        "Tier 2 - Prospect":  PatternFill("solid", fgColor="FEFCE8"),
        "Tier 3 - Potential": PatternFill("solid", fgColor="F0FDF4"),
        "Tier 4 - Unlikely":  PatternFill("solid", fgColor="F8F7F4"),
    }

    # ── Sheet 1: All New RFPs ──
    ws1 = wb.active
    ws1.title = "New RFPs This Week"

    cols = [
        ("Tier", "priority", 15),
        ("Score", "score", 8),
        ("Title", "name", 50),
        ("Issuing Org", "end_client", 25),
        ("Portal", "source_portal", 15),
        ("Closing", "closing_date", 12),
        ("Office", "office", 15),
        ("Country", "country", 8),
        ("Sector", "sector", 20),
        ("Contract Type", "contract_type", 15),
        ("Budget (USD)", "project_value_usd", 15),
        ("Link", "details_url", 40),
    ]

    # Headers
    for col_idx, (header, _, width) in enumerate(cols, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws1.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, (_, field, _) in enumerate(cols, 1):
            val = row.get(field, "")
            if val is None:
                val = ""
            if field == "project_value_usd" and val:
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    pass
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.border = border

            # Color tier column
            if field == "priority" and val in tier_colors:
                cell.fill = tier_colors[val]

            # Format date columns
            if field in ("closing_date", "questions_deadline") and val:
                if hasattr(val, "strftime"):
                    cell.value = val.strftime("%Y-%m-%d")

            # Format as currency
            if field == "project_value_usd" and isinstance(val, (int, float)):
                cell.number_format = "$#,##0"

    ws1.auto_filter.ref = ws1.dimensions
    ws1.freeze_panes = "A2"

    # ── Sheet 2: By Portal ──
    ws2 = wb.create_sheet("By Portal")
    portal_stats = df.groupby("source_portal").agg(
        count=("solicitation_id", "count"),
        avg_score=("score", "mean"),
    ).sort_values("count", ascending=False)

    portal_headers = ["Portal", "Count", "Avg Score"]
    for col_idx, header in enumerate(portal_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        ws2.column_dimensions[get_column_letter(col_idx)].width = 20

    for row_idx, (portal, stats) in enumerate(portal_stats.iterrows(), 2):
        ws2.cell(row=row_idx, column=1, value=portal).font = cell_font
        ws2.cell(row=row_idx, column=2, value=int(stats["count"])).font = cell_font
        avg = stats["avg_score"]
        ws2.cell(row=row_idx, column=3, value=round(avg, 1) if avg else 0).font = cell_font

    # ── Sheet 3: Summary ──
    ws3 = wb.create_sheet("Summary")
    summary_data = [
        ("EITACIES RFP GenAI — Weekly Report", ""),
        ("Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("", ""),
        ("Total RFPs This Week", len(df)),
        ("Portals Covered", df["source_portal"].nunique()),
        ("Avg Score", round(df["score"].mean(), 1) if len(df) > 0 else 0),
        ("", ""),
        ("Tier Breakdown", ""),
    ]

    tier_counts = df["priority"].value_counts()
    for tier in ["Strategic", "Tier 1 - Pursue", "Tier 2 - Prospect", "Tier 3 - Potential", "Tier 4 - Unlikely"]:
        count = tier_counts.get(tier, 0)
        summary_data.append((f"  {tier}", count))

    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws3.cell(row=row_idx, column=1, value=label).font = cell_font
        ws3.cell(row=row_idx, column=2, value=value).font = cell_font
    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 20

    # Bold title
    ws3.cell(row=1, column=1).font = Font(size=14, bold=True, name="Calibri")

    return wb


def main():
    """Generate the weekly Excel report."""
    print("EITACIES RFP GenAI — Excel Report Generator")
    print("=" * 50)

    # Pull data
    df = pull_records()
    print(f"Records this week: {len(df)}")

    if len(df) == 0:
        print("No records to report. Injecting sample data for report...")
        # Create minimal sample for report
        import pandas as pd
        df = pd.DataFrame([{
            "solicitation_id": "SAMPLE-001",
            "name": "Sample RFP — AI/ML Platform",
            "source_portal": "SAM.gov",
            "end_client": "Federal Agency",
            "priority": "Tier 2 - Prospect",
            "score": 65,
            "closing_date": None,
            "country": "US",
            "office": "",
            "sector": "Technology",
            "contract_type": "",
            "project_value_usd": None,
            "project_value_cad": None,
            "details_url": "",
            "description": "",
            "notes": "",
            "flags": [],
            "red_flags": [],
            "status": "new",
        }])

    # Generate Excel
    wb = generate_excel(df)

    # Save
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"EITACIES_RFPs_Week_{datetime.utcnow().strftime('%Y-W%W')}.xlsx"
    output_path = OUTPUT_DIR / filename
    wb.save(output_path)

    print(f"Excel report saved: {output_path}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Rows in Sheet 1: {wb.active.max_row}")
    print("Done!")

    return str(output_path)


if __name__ == "__main__":
    main()
